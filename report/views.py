from django.http.response import HttpResponse
from django.shortcuts import render

from django.http import HttpResponse
from django.contrib.auth.models import User
from io import BytesIO
from openpyxl.workbook.workbook import Workbook
import xlwt
from openpyxl import load_workbook
from report.form import UploadData
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import string

from report.models import CostCenter, ReportItem

def downloadSample(request):
  costCenter = request.user.reporter.costCenter
  segment = request.user.reporter.costCenter.segment
  costCenterCategory = costCenter.costCenterCategory
  # return HttpResponse(costCenterCategory.report_item_set.all())

  wb = Workbook()
  ws = wb.active
  ws.title = "Upload budget"
  header = ["STT", "Segment",	"Segment Name", "Cost center", "Cost center name", "Item code", "Item Name", "Month", "Amount"]
  ws.append(header)

  for alphabet in list(string.ascii_uppercase):
    font = Font(bold=True)
    cell = ws[alphabet + "1"]
    cell.font = font
  rowIndex = 2
  for reportItem in ReportItem.objects.filter(costCenterCategory = costCenterCategory, parent = None):
    row = [rowIndex - 1, segment.segmentCode, segment.segmentName, costCenter.costCenterCode, costCenter.costCenterName, reportItem.reportItemCode, reportItem.reportItemName]
    ws.append(row)
    rowIndex = rowIndex + 1
    for childReportItem in ReportItem.objects.filter(costCenterCategory = costCenterCategory, parent = reportItem):
      row = [rowIndex - 1, segment.segmentCode, segment.segmentName, costCenter.costCenterCode, costCenter.costCenterName, childReportItem.reportItemCode, childReportItem.reportItemName]
      ws.append(row)
      alignment = Alignment(
        indent=5
      )
      for alphabet in list(string.ascii_uppercase):
        font = Font(bold=True)
      cell = ws[alphabet + str(rowIndex)]
      cell.alignment = alignment
      rowIndex = rowIndex + 1


  response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
  return response

def sample(request):
  f = UploadData()
  if request.method == "POST":
    # print(request.FILES)
    f = UploadData(request.POST, request.FILES)
    if f.is_valid():
      # print(request.FILES['file'].file)
      # wb = load_workbook(filename=request.FILES['file'].file)
      file_in_memory = request.FILES['file'].read()
      wb = load_workbook(filename=BytesIO(file_in_memory))
      ws = wb.active
      # print(f.cleaned_data)
      segment = f.cleaned_data.get("segment")
      print(segment)
      print(ws['A1'])
      print(ws['A1'].value)
      print(ws['A1'].is_date)
      print(ws['A1'].data_type)
      print(ws['A1'].row)

  return render(request, 'report/sample.html', {"f": f})

